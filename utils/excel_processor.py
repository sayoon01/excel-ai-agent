"""Excel-specific processing utilities (header detection, column classification, merge).
Adapted from cowork-llm-lab."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd


# ── Header detection ───────────────────────────────────────────────────────────

def detect_header_row(df_raw: pd.DataFrame, max_scan: int = 15) -> int:
    """Return the 0-based row index most likely to be the header."""
    best_row, best_score = 0, -1
    for i in range(min(max_scan, len(df_raw))):
        row = df_raw.iloc[i]
        non_null = row.count()
        if non_null == 0:
            continue
        non_numeric = sum(
            1 for v in row if isinstance(v, str) and not _is_numeric_str(v)
        )
        score = (non_null / len(row)) * 0.6 + (non_numeric / max(non_null, 1)) * 0.4
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _is_numeric_str(s: str) -> bool:
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


def read_with_header(path: Path) -> pd.DataFrame:
    """Read Excel/CSV with auto header detection."""
    raw = (
        pd.read_csv(path, header=None)
        if path.suffix.lower() == ".csv"
        else pd.read_excel(path, header=None)
    )
    header_row = detect_header_row(raw)
    df = (
        pd.read_csv(path, header=header_row)
        if path.suffix.lower() == ".csv"
        else pd.read_excel(path, header=header_row)
    )
    return df


# ── Column classification ──────────────────────────────────────────────────────

def classify_columns(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    """Return (key_columns, numeric_columns)."""
    key_cols, num_cols = [], []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            num_cols.append(col)
        else:
            key_cols.append(col)
    return key_cols, num_cols


# ── Multi-file merge ───────────────────────────────────────────────────────────

def merge_dataframes(
    dfs: list[pd.DataFrame],
    key_cols: list[str],
    num_cols: list[str],
    agg: str = "sum",
) -> pd.DataFrame:
    """Merge multiple DataFrames on key columns and aggregate numeric columns."""
    combined = pd.concat(dfs, ignore_index=True)
    if not key_cols:
        return combined
    agg_fns = {col: agg for col in num_cols if col in combined.columns}
    if agg_fns:
        return combined.groupby(key_cols, as_index=False).agg(agg_fns)
    return combined.drop_duplicates(subset=key_cols)


def sanitize_sheet_name(name: str) -> str:
    """Replace characters that Excel forbids in sheet names."""
    return re.sub(r"[\\/:*?\[\]]", "_", name)[:31]


# ── 실제 데이터 입력 범위 탐지 ────────────────────────────────────────────────

@dataclass
class UsedRange:
    first_row: int      # 데이터 시작 행 (1-based, 헤더 포함)
    last_row: int       # 데이터 마지막 행 (1-based)
    first_col: int      # 데이터 시작 열 (1-based)
    last_col: int       # 데이터 마지막 열 (1-based)
    data_rows: int      # 값이 1개 이상 있는 행 수
    data_cols: int      # 값이 1개 이상 있는 열 수
    filled_cells: int   # 실제 값이 있는 셀 수
    total_cells: int    # 범위 내 전체 셀 수
    density: float      # filled / total (0.0 ~ 1.0)
    sheet_name: str     # 탐지한 시트 이름 (xlsx만)


def get_used_range(path: Path) -> UsedRange | None:
    """파일에서 실제 데이터가 입력된 셀 범위를 탐지."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _used_range_excel(path)
    elif suffix == ".csv":
        return _used_range_csv(path)
    return None


def _used_range_excel(path: Path) -> UsedRange | None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title

        # openpyxl이 계산한 사용 범위
        min_r, max_r = ws.min_row, ws.max_row
        min_c, max_c = ws.min_column, ws.max_column
        if min_r is None:
            return None

        # 실제 값 있는 셀 수 계산
        filled = 0
        row_has_data: set[int] = set()
        col_has_data: set[int] = set()

        for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                 min_col=min_c, max_col=max_c):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    filled += 1
                    row_has_data.add(cell.row)
                    col_has_data.add(cell.column)

        wb.close()
        total = (max_r - min_r + 1) * (max_c - min_c + 1)
        return UsedRange(
            first_row=min_r, last_row=max_r,
            first_col=min_c, last_col=max_c,
            data_rows=len(row_has_data),
            data_cols=len(col_has_data),
            filled_cells=filled,
            total_cells=total,
            density=round(filled / total, 3) if total > 0 else 0.0,
            sheet_name=sheet_name,
        )
    except Exception:
        return None


def _used_range_csv(path: Path) -> UsedRange | None:
    try:
        df = pd.read_csv(path, header=None)
        if df.empty:
            return None

        # 값이 있는 행/열 인덱스 찾기
        row_mask = df.notna().any(axis=1)
        col_mask = df.notna().any(axis=0)

        first_r = int(row_mask.idxmax()) + 1
        last_r  = int(row_mask[::-1].idxmax()) + 1
        first_c = int(col_mask.idxmax()) + 1
        last_c  = int(col_mask[::-1].idxmax()) + 1

        filled = int(df.notna().sum().sum())
        total  = (last_r - first_r + 1) * (last_c - first_c + 1)
        return UsedRange(
            first_row=first_r, last_row=last_r,
            first_col=first_c, last_col=last_c,
            data_rows=int(row_mask.sum()),
            data_cols=int(col_mask.sum()),
            filled_cells=filled,
            total_cells=total,
            density=round(filled / total, 3) if total > 0 else 0.0,
            sheet_name="(CSV)",
        )
    except Exception:
        return None


def format_used_range(ur: UsedRange) -> str:
    """UsedRange를 사람이 읽기 쉬운 문자열로 변환."""
    return (
        f"범위: {ur.first_row}행~{ur.last_row}행 / "
        f"{ur.first_col}열~{ur.last_col}열  |  "
        f"데이터: {ur.data_rows}행 × {ur.data_cols}열  |  "
        f"채워진 셀: {ur.filled_cells}/{ur.total_cells} "
        f"({ur.density*100:.1f}%)"
    )


# ── Sheet summary builder ──────────────────────────────────────────────────────

def describe_dataframe(df: pd.DataFrame, fname: str) -> str:
    """Return a human-readable summary string for LLM context."""
    key_cols, num_cols = classify_columns(df)
    lines = [
        f"파일: {fname}  ({len(df)}행 × {len(df.columns)}열)",
        f"  키 컬럼 : {', '.join(key_cols) or '없음'}",
        f"  숫자 컬럼: {', '.join(num_cols) or '없음'}",
        f"  샘플:",
    ]
    lines.append(df.head(3).to_string(index=False))
    return "\n".join(lines)
