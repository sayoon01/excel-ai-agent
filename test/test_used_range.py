"""used_range 기능 검증 스크립트.

실행: source venv/bin/activate && python test_used_range.py
"""
from pathlib import Path
import pandas as pd
import openpyxl

from utils.excel_processor import get_used_range, format_used_range

TMP = Path("uploads/_test_range.xlsx")

# ── 테스트 파일 생성 ──────────────────────────────────────────────────────────
# 실제 데이터는 B3:D7 (빈 행/열이 앞뒤에 있는 상황 시뮬레이션)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# B3에서 시작하는 데이터 (A열, 1~2행은 비워둠)
headers = ["이름", "매출", "지역"]
data = [
    ["홍길동", 150, "서울"],
    ["김철수", None, "부산"],   # 결측치 포함
    ["이영희", 200, "대구"],
    ["박민수", 80,  "서울"],
]

for col_idx, h in enumerate(headers, start=2):   # B=2부터
    ws.cell(row=3, column=col_idx, value=h)
for row_idx, row in enumerate(data, start=4):
    for col_idx, val in enumerate(row, start=2):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(TMP)
print(f"테스트 파일 생성: {TMP}")
print(f"  설계 범위: 3행~7행, B열(2)~D열(4)\n")

# ── 탐지 실행 ─────────────────────────────────────────────────────────────────
ur = get_used_range(TMP)
assert ur is not None, "탐지 실패"

print("=== 탐지 결과 ===")
print(f"  시트명      : {ur.sheet_name}")
print(f"  시작 위치   : {ur.first_row}행 / {ur.first_col}열")
print(f"  끝 위치     : {ur.last_row}행 / {ur.last_col}열")
print(f"  데이터 행수  : {ur.data_rows}")
print(f"  데이터 열수  : {ur.data_cols}")
print(f"  채워진 셀    : {ur.filled_cells} / {ur.total_cells}")
print(f"  밀도         : {ur.density*100:.1f}%")
print()
print("=== format_used_range ===")
print(" ", format_used_range(ur))

# ── 검증 ─────────────────────────────────────────────────────────────────────
assert ur.first_row == 3,  f"시작행 오류: {ur.first_row}"
assert ur.last_row  == 7,  f"끝행 오류: {ur.last_row}"
assert ur.first_col == 2,  f"시작열 오류: {ur.first_col}"
assert ur.last_col  == 4,  f"끝열 오류: {ur.last_col}"
assert ur.data_rows == 5,  f"데이터행 오류: {ur.data_rows}"   # 헤더 포함 5행
assert ur.data_cols == 3,  f"데이터열 오류: {ur.data_cols}"
# 결측치 1개이므로 filled = 5*3 - 1 = 14
assert ur.filled_cells == 14, f"filled 오류: {ur.filled_cells}"

print("\n모든 검증 통과")

# ── CSV 검증 ──────────────────────────────────────────────────────────────────
csv_path = Path("uploads/_test_range.csv")
df = pd.DataFrame(data, columns=headers)
df.to_csv(csv_path, index=False)

ur_csv = get_used_range(csv_path)
assert ur_csv is not None
print(f"\n=== CSV 탐지 결과 ===")
print(" ", format_used_range(ur_csv))

# 정리
TMP.unlink()
csv_path.unlink()
print("\n테스트 파일 삭제 완료")
