"""Safe sandboxed Python/pandas code execution for LLM-generated code."""
from __future__ import annotations

import ast
import io
import re
import signal
import traceback
import uuid
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from core.llm.llm_client import LLMClient

from functools import reduce

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from services.file_manager import RESULT_DIR, list_files, read_file

_CORRECTION_SYSTEM = (
    "Python/pandas 코드 디버거입니다. 오류를 수정한 코드 블록만 반환하세요. 설명 불필요.\n"
    "환경: files(dict), pd, np, plt 주입됨. import 문 사용 불가. "
    "최종 결과는 반드시 result 변수에 저장."
)

_CODE_BLOCK_RE = re.compile(r"```python\s*\n(.*?)```", re.DOTALL)

# pd, np, plt, matplotlib, functools는 namespace에 이미 주입 — import 문만 조용히 제거
_PRE_INJECTED_IMPORT = re.compile(
    r"^[ \t]*(?:import (?:pandas|numpy|matplotlib|functools)(?:\.\w+)*(?:\s+as\s+\w+)?|from (?:pandas|numpy|matplotlib|functools)(?:\.\w+)*\b[^\n]*)[ \t]*(?:\n|$)",
    re.MULTILINE,
)


def _strip_preinjected_imports(code: str) -> str:
    return _PRE_INJECTED_IMPORT.sub("", code)


BLOCKED_MODULES = frozenset({
    "os", "subprocess", "sys", "shutil", "importlib",
    "socket", "http", "urllib", "requests", "httpx",
    "pathlib", "glob", "pickle", "shelve", "marshal",
    "ctypes", "multiprocessing", "threading",
    "signal", "atexit", "code", "codeop", "compileall",
})

BLOCKED_BUILTINS = frozenset({
    "exec", "eval", "compile", "__import__",
    "open", "input", "breakpoint",
    "globals", "locals", "vars",
    "getattr", "setattr", "delattr",
    "memoryview",
})


@dataclass
class ExecutionResult:
    success: bool
    output: str = ""
    error: str = ""
    result_df: pd.DataFrame | None = None
    result_type: str = ""      # "dataframe" | "number" | "string" | "plot"
    result_value: object = None
    is_corrected: bool = False
    correction_attempts: int = 0
    saved_files: list[str] = field(default_factory=list)


def _validate_code(code: str) -> list[str]:
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return [f"문법 오류: {e}"]

    violations: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            # pd/np는 namespace에 이미 주입됨. __import__는 샌드박스에서 차단됨.
            if isinstance(node, ast.Import):
                names = [alias.name for alias in node.names]
            else:
                mod = node.module or ""
                names = [mod] if mod else [alias.name for alias in node.names]
            for name in names:
                root = (name or "").split(".")[0]
                if root in BLOCKED_MODULES:
                    violations.append(f"허용되지 않는 import: {name}")
                else:
                    violations.append(
                        f"import 문 사용 불가: {name} "
                        "(pd, np는 이미 사용 가능 — import 없이 바로 쓰세요)"
                    )
        elif isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in BLOCKED_BUILTINS:
                violations.append(f"차단된 내장 함수: {node.func.id}")
    return violations


def _make_safe_builtins() -> dict:
    import builtins
    safe = {}
    for name in dir(builtins):
        if name not in BLOCKED_BUILTINS and not name.startswith("_"):
            safe[name] = getattr(builtins, name)
    safe["__build_class__"] = builtins.__build_class__
    safe["__name__"] = "__main__"
    return safe


def _apply_xlsx_formatting(dest: Path, df: pd.DataFrame) -> None:
    """openpyxl로 헤더 볼드, 숫자 천단위 포맷, 컬럼 너비 자동 조정 적용."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(dest)
        ws = wb.active

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        num_fmt = '#,##0.##'

        numeric_col_indices = {
            i + 1
            for i, col in enumerate(df.columns)
            if pd.api.types.is_numeric_dtype(df[col])
        }

        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx, cell in enumerate(ws[col_letter]):
                if col_idx in numeric_col_indices and row_idx > 0:
                    cell.number_format = num_fmt
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

        wb.save(dest)
    except Exception:
        pass


def _make_save_fn(saved_files: list[str], namespace: dict):
    def save(filename: str, df: pd.DataFrame | None = None):
        if df is None:
            df = namespace.get("result")
        if df is None:
            raise ValueError("저장할 DataFrame이 없습니다. result에 할당하거나 df 인자를 전달하세요.")
        if not isinstance(df, pd.DataFrame):
            raise TypeError(f"DataFrame이 필요합니다. {type(df).__name__} 전달됨.")
        name = Path(filename).name
        if "/" in filename or "\\" in filename or ".." in filename:
            raise ValueError("유효하지 않은 파일명입니다.")
        suffix = Path(name).suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            raise ValueError(f"지원하지 않는 형식: {suffix}. .xlsx 또는 .csv를 사용하세요.")
        dest = RESULT_DIR / name
        if suffix == ".csv":
            df.to_csv(dest, index=False)
        else:
            df.to_excel(dest, index=False)
            _apply_xlsx_formatting(dest, df)
        saved_files.append(name)
    return save


class _Timeout:
    def __init__(self, seconds: int):
        self.seconds = seconds
        self._old = None

    def __enter__(self):
        try:
            self._old = signal.signal(signal.SIGALRM, self._handler)
            signal.alarm(self.seconds)
        except (ValueError, OSError):
            pass
        return self

    def __exit__(self, *_):
        try:
            signal.alarm(0)
            if self._old is not None:
                signal.signal(signal.SIGALRM, self._old)
        except (ValueError, OSError):
            pass

    @staticmethod
    def _handler(signum, frame):
        raise TimeoutError("코드 실행 시간이 초과되었습니다.")


def execute(
    code: str,
    timeout_seconds: int = 30,
    last_result: pd.DataFrame | None = None,
    selected_sheets: dict[str, str] | None = None,
    selected_files: list[str] | None = None,
) -> ExecutionResult:
    code = _strip_preinjected_imports(code)
    violations = _validate_code(code)
    if violations:
        return ExecutionResult(
            success=False,
            error="코드 검증 실패:\n" + "\n".join(f"  - {v}" for v in violations),
        )

    saved_files: list[str] = []
    _sheets = selected_sheets or {}
    _file_filter = set(selected_files) if selected_files else None
    all_fnames = [f for f in list_files() if _file_filter is None or f in _file_filter]
    files = {fname: read_file(fname, sheet_name=_sheets.get(fname)) for fname in all_fnames}
    files = {k: v for k, v in files.items() if v is not None}

    namespace: dict = {
        "files": files,
        "last_result": last_result,
        "pd": pd,
        "np": np,
        "plt": plt,
        "matplotlib": matplotlib,
        "reduce": reduce,
        "result": None,
        "__builtins__": _make_safe_builtins(),
    }
    namespace["save"] = _make_save_fn(saved_files, namespace)

    stdout_capture = io.StringIO()
    namespace["print"] = lambda *args, **kwargs: print(
        *args, **kwargs, file=stdout_capture
    )

    try:
        with _Timeout(timeout_seconds):
            exec(compile(code, "<llm_generated>", "exec"), namespace)  # noqa: S102
    except TimeoutError as e:
        return ExecutionResult(
            success=False,
            output=stdout_capture.getvalue(),
            error=str(e),
        )
    except Exception:
        return ExecutionResult(
            success=False,
            output=stdout_capture.getvalue(),
            error=traceback.format_exc(),
        )

    result_raw = namespace.get("result")
    result_type = ""
    result_value: object = result_raw
    result_df: pd.DataFrame | None = None

    if isinstance(result_raw, dict) and "type" in result_raw and "value" in result_raw:
        result_type = str(result_raw["type"])
        result_value = result_raw["value"]
        if result_type == "dataframe" and isinstance(result_value, pd.DataFrame):
            result_df = result_value
        elif result_type == "plot" and hasattr(result_value, "savefig"):
            chart_path = RESULT_DIR / f"chart_{uuid.uuid4().hex[:8]}.png"
            result_value.savefig(chart_path, dpi=150, bbox_inches="tight")
            plt.close(result_value)
            result_value = str(chart_path)
    elif isinstance(result_raw, pd.DataFrame):
        result_type = "dataframe"
        result_df = result_raw
    elif isinstance(result_raw, (int, float)) and result_raw is not None:
        result_type = "number"
    elif isinstance(result_raw, str):
        result_type = "string"

    # save() 없이 result만 있으면 results/에 자동 저장 (사이드바·다운로드 버튼 연동)
    if result_df is not None and not saved_files:
        auto_name = f"result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        auto_dest = RESULT_DIR / auto_name
        result_df.to_excel(auto_dest, index=False)
        _apply_xlsx_formatting(auto_dest, result_df)
        saved_files.append(auto_name)

    return ExecutionResult(
        success=True,
        output=stdout_capture.getvalue(),
        result_df=result_df,
        result_type=result_type,
        result_value=result_value,
        saved_files=saved_files,
    )


def _build_file_schema(
    selected_files: list[str] | None,
    selected_sheets: dict[str, str] | None,
) -> str:
    """사용 가능한 파일의 컬럼 목록 문자열 — 수정 프롬프트에 포함."""
    _sheets = selected_sheets or {}
    _filter = set(selected_files) if selected_files else None
    lines: list[str] = []
    for fname in list_files():
        if _filter and fname not in _filter:
            continue
        df = read_file(fname, sheet_name=_sheets.get(fname))
        if df is not None:
            cols = ", ".join(f'"{c}"' for c in df.columns[:25])
            lines.append(f"- {fname} ({len(df)}행): [{cols}]")
    return "\n".join(lines) if lines else "(파일 없음)"


def execute_with_retry(
    code: str,
    last_result: pd.DataFrame | None = None,
    client: "LLMClient | None" = None,
    original_question: str = "",
    max_attempts: int = 3,
    selected_sheets: dict[str, str] | None = None,
    selected_files: list[str] | None = None,
) -> ExecutionResult:
    """코드를 실행하고 실패 시 LLM에게 수정을 요청해 재시도한다.

    client가 None이면 단순 execute()와 동일하게 동작한다.
    개선: 수정 프롬프트에 실제 파일 컬럼 정보를 포함해 컬럼명 오류 자동 수정.
    """
    _exec_kw = {"selected_sheets": selected_sheets, "selected_files": selected_files}
    result = execute(code, last_result=last_result, **_exec_kw)
    if result.success or client is None:
        return result

    # 타임아웃·보안 위반은 재시도해도 해결 불가
    _no_retry_signals = ("시간이 초과", "TimeoutError", "코드 검증 실패", "허용되지 않는")
    if any(s in result.error for s in _no_retry_signals):
        return result

    # 수정 프롬프트에 포함할 실제 파일 스키마
    _schema = _build_file_schema(selected_files, selected_sheets)

    current_code = code
    attempts = 0
    for _ in range(max_attempts - 1):
        # KeyError면 잘못 사용한 컬럼명을 에러 메시지에서 추출해 명시
        _keyerror_hint = ""
        _ke_match = re.search(r"KeyError:\s*['\"]?([^'\"\n]+)['\"]?", result.error)
        if _ke_match:
            _keyerror_hint = (
                f"\n[참고] 존재하지 않는 컬럼 '{_ke_match.group(1).strip()}'을 사용했습니다. "
                f"아래 실제 컬럼명 목록에서 올바른 이름을 찾아 수정하세요.\n"
            )

        correction_msg = (
            f"원래 질문: {original_question}\n\n"
            f"실행한 코드:\n```python\n{current_code}\n```\n\n"
            f"오류 메시지:\n{result.error}{_keyerror_hint}\n\n"
            f"사용 가능한 파일과 실제 컬럼명:\n{_schema}\n\n"
            f"위 컬럼명을 참고해 오류를 수정한 코드를 제공해주세요."
        )
        try:
            raw = "".join(client.chat_stream(
                [{"role": "user", "content": correction_msg}],
                _CORRECTION_SYSTEM,
            ))
        except Exception:
            break

        codes = _CODE_BLOCK_RE.findall(raw)
        if not codes:
            break

        current_code = codes[0]
        attempts += 1
        result = execute(current_code, last_result=last_result, **_exec_kw)
        if result.success:
            result.is_corrected = True
            result.correction_attempts = attempts
            return result

    result.correction_attempts = attempts
    return result
