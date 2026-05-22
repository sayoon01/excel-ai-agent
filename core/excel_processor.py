"""Excel-specific processing utilities (header detection, column classification, merge).
Adapted from cowork-llm-lab."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd


# ── Header detection ───────────────────────────────────────────────────────────

def detect_header_row(df_raw: pd.DataFrame, max_scan: int = 15) -> int:
    """Return the 0-based row index most likely to be the header."""
    best_row, best_score = 0, -1
    for i in range(min(max_scan, len(df_raw))):
        row = df_raw.iloc[i]
        non_null = row.count()
        if non_null == 0:
            continue
        non_numeric = sum(
            1 for v in row if isinstance(v, str) and not _is_numeric_str(v)
        )
        score = (non_null / len(row)) * 0.6 + (non_numeric / max(non_null, 1)) * 0.4
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _is_numeric_str(s: str) -> bool:
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


def read_with_header(path: Path) -> pd.DataFrame:
    """Read Excel/CSV with auto header detection."""
    raw = (
        pd.read_csv(path, header=None)
        if path.suffix.lower() == ".csv"
        else pd.read_excel(path, header=None)
    )
    header_row = detect_header_row(raw)
    df = (
        pd.read_csv(path, header=header_row)
        if path.suffix.lower() == ".csv"
        else pd.read_excel(path, header=header_row)
    )
    return df


def read_excel_smart(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    """엑셀 파일 스마트 읽기.

    1. 헤더 행 자동 탐지 (빈 행·제목 행 건너뜀)
    2. 2행 헤더(멀티레벨) 감지 → '상위_하위' 형식으로 평탄화
    3. 병합 헤더의 Unnamed 컬럼 → 앞 컬럼명 기반으로 이름 부여
    4. 데이터 영역 텍스트 컬럼 병합셀(NaN) → ffill
    5. 숫자처럼 생긴 문자열(쉼표 포함) → numeric 변환
    6. 소수점 없는 float 컬럼 → Int64 다운캐스트
    """
    read_kw = {"sheet_name": sheet_name} if sheet_name is not None else {}
    try:
        raw = pd.read_excel(path, header=None, **read_kw)
        if raw.empty:
            return raw

        header_row = detect_header_row(raw)

        # ── 멀티레벨 헤더 감지 ─────────────────────────────────────────────
        use_multi = False
        if header_row + 1 < len(raw):
            next_row = raw.iloc[header_row + 1]
            non_null = next_row.count()
            non_numeric = sum(
                1 for v in next_row
                if isinstance(v, str) and not _is_numeric_str(v)
            )
            if non_null >= 3 and non_numeric / max(non_null, 1) >= 0.5:
                use_multi = True

        if use_multi:
            df = pd.read_excel(path, header=[header_row, header_row + 1], **read_kw)
            df = _flatten_multiindex(df)
        else:
            df = pd.read_excel(path, header=header_row, **read_kw)
            df = _fix_unnamed_cols(df, raw.iloc[header_row])

        # ── 텍스트 컬럼 병합셀(NaN) → ffill ───────────────────────────────
        text_cols = df.select_dtypes(include=["object"]).columns
        if len(text_cols):
            df[text_cols] = df[text_cols].ffill()

        # ── 숫자처럼 생긴 문자열 → numeric ────────────────────────────────
        df = _coerce_numeric_cols(df)

        # ── 소수점 없는 float → Int64 ──────────────────────────────────────
        df = _downcast_floats(df)

        return df
    except Exception:
        return pd.read_excel(path, **read_kw)


# ── read_excel_smart 내부 헬퍼 ────────────────────────────────────────────────

def _flatten_multiindex(df: pd.DataFrame) -> pd.DataFrame:
    """MultiIndex 컬럼을 '상위_하위' 단일 레벨로 평탄화."""
    new_cols = []
    for col in df.columns:
        if isinstance(col, tuple):
            parts = [str(c).strip() for c in col if not str(c).startswith("Unnamed")]
            new_cols.append("_".join(parts) if parts else f"col_{len(new_cols)}")
        else:
            new_cols.append(str(col))
    # 중복 컬럼명 처리
    counts: dict[str, int] = {}
    result = []
    for name in new_cols:
        if name in counts:
            counts[name] += 1
            result.append(f"{name}_{counts[name]}")
        else:
            counts[name] = 0
            result.append(name)
    df.columns = result
    return df


def _fix_unnamed_cols(df: pd.DataFrame, header_series: pd.Series) -> pd.DataFrame:
    """병합 헤더로 생긴 Unnamed 컬럼에 '이전컬럼명_N' 형식으로 이름 부여."""
    cols = list(df.columns)
    raw_vals = header_series.tolist()
    last_named: str | None = None
    suffix_cnt: dict[str, int] = {}

    for i, (col, raw_val) in enumerate(zip(cols, raw_vals)):
        col_str = str(col)
        if col_str.startswith("Unnamed:"):
            raw_str = "" if pd.isna(raw_val) else str(raw_val).strip()
            if raw_str and not raw_str.startswith("Unnamed"):
                # 원본 값이 있으면 그걸 사용
                last_named = raw_str
                suffix_cnt[last_named] = suffix_cnt.get(last_named, 0)
                cols[i] = last_named
            elif last_named:
                suffix_cnt[last_named] = suffix_cnt.get(last_named, 0) + 1
                cols[i] = f"{last_named}_{suffix_cnt[last_named]}"
        else:
            last_named = col_str

    df.columns = cols
    return df


def _coerce_numeric_cols(df: pd.DataFrame) -> pd.DataFrame:
    """object 컬럼에서 50% 이상이 숫자(쉼표 포함)이면 numeric으로 변환."""
    for col in df.select_dtypes(include=["object"]).columns:
        cleaned = (
            df[col].astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
            .replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
        )
        numeric = pd.to_numeric(cleaned, errors="coerce")
        if numeric.notna().mean() >= 0.5:
            df[col] = numeric
    return df


def _downcast_floats(df: pd.DataFrame) -> pd.DataFrame:
    """소수점이 없는 float64 컬럼 → pandas Int64 (nullable integer)."""
    for col in df.select_dtypes(include=["float64"]).columns:
        non_null = df[col].dropna()
        if len(non_null) and (non_null % 1 == 0).all():
            df[col] = df[col].astype("Int64")
    return df


# ── Column classification ──────────────────────────────────────────────────────

def classify_columns(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    """Return (key_columns, numeric_columns)."""
    key_cols, num_cols = [], []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            num_cols.append(col)
        else:
            key_cols.append(col)
    return key_cols, num_cols


# ── Multi-file merge ───────────────────────────────────────────────────────────

def merge_dataframes(
    dfs: list[pd.DataFrame],
    key_cols: list[str],
    num_cols: list[str],
    agg: str = "sum",
) -> pd.DataFrame:
    """Merge multiple DataFrames on key columns and aggregate numeric columns."""
    combined = pd.concat(dfs, ignore_index=True)
    if not key_cols:
        return combined
    agg_fns = {col: agg for col in num_cols if col in combined.columns}
    if agg_fns:
        return combined.groupby(key_cols, as_index=False).agg(agg_fns)
    return combined.drop_duplicates(subset=key_cols)


def sanitize_sheet_name(name: str) -> str:
    """Replace characters that Excel forbids in sheet names."""
    return re.sub(r"[\\/:*?\[\]]", "_", name)[:31]


# ── 실제 데이터 입력 범위 탐지 ────────────────────────────────────────────────

@dataclass
class UsedRange:
    first_row: int      # 데이터 시작 행 (1-based, 헤더 포함)
    last_row: int       # 데이터 마지막 행 (1-based)
    first_col: int      # 데이터 시작 열 (1-based)
    last_col: int       # 데이터 마지막 열 (1-based)
    data_rows: int      # 값이 1개 이상 있는 행 수
    data_cols: int      # 값이 1개 이상 있는 열 수
    filled_cells: int   # 실제 값이 있는 셀 수
    total_cells: int    # 범위 내 전체 셀 수
    density: float      # filled / total (0.0 ~ 1.0)
    sheet_name: str     # 탐지한 시트 이름 (xlsx만)


def get_used_range(path: Path) -> UsedRange | None:
    """파일에서 실제 데이터가 입력된 셀 범위를 탐지."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _used_range_excel(path)
    elif suffix == ".csv":
        return _used_range_csv(path)
    return None


def _used_range_excel(path: Path) -> UsedRange | None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title

        # openpyxl이 계산한 사용 범위
        min_r, max_r = ws.min_row, ws.max_row
        min_c, max_c = ws.min_column, ws.max_column
        if min_r is None:
            return None

        # 실제 값 있는 셀 수 계산
        filled = 0
        row_has_data: set[int] = set()
        col_has_data: set[int] = set()

        for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                 min_col=min_c, max_col=max_c):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    filled += 1
                    row_has_data.add(cell.row)
                    col_has_data.add(cell.column)

        wb.close()
        total = (max_r - min_r + 1) * (max_c - min_c + 1)
        return UsedRange(
            first_row=min_r, last_row=max_r,
            first_col=min_c, last_col=max_c,
            data_rows=len(row_has_data),
            data_cols=len(col_has_data),
            filled_cells=filled,
            total_cells=total,
            density=round(filled / total, 3) if total > 0 else 0.0,
            sheet_name=sheet_name,
        )
    except Exception:
        return None


def _used_range_csv(path: Path) -> UsedRange | None:
    try:
        df = pd.read_csv(path, header=None)
        if df.empty:
            return None

        # 값이 있는 행/열 인덱스 찾기
        row_mask = df.notna().any(axis=1)
        col_mask = df.notna().any(axis=0)

        first_r = int(row_mask.idxmax()) + 1
        last_r  = int(row_mask[::-1].idxmax()) + 1
        first_c = int(col_mask.idxmax()) + 1
        last_c  = int(col_mask[::-1].idxmax()) + 1

        filled = int(df.notna().sum().sum())
        total  = (last_r - first_r + 1) * (last_c - first_c + 1)
        return UsedRange(
            first_row=first_r, last_row=last_r,
            first_col=first_c, last_col=last_c,
            data_rows=int(row_mask.sum()),
            data_cols=int(col_mask.sum()),
            filled_cells=filled,
            total_cells=total,
            density=round(filled / total, 3) if total > 0 else 0.0,
            sheet_name="(CSV)",
        )
    except Exception:
        return None


def format_used_range(ur: UsedRange) -> str:
    """UsedRange를 사람이 읽기 쉬운 문자열로 변환."""
    return (
        f"범위: {ur.first_row}행~{ur.last_row}행 / "
        f"{ur.first_col}열~{ur.last_col}열  |  "
        f"데이터: {ur.data_rows}행 × {ur.data_cols}열  |  "
        f"채워진 셀: {ur.filled_cells}/{ur.total_cells} "
        f"({ur.density*100:.1f}%)"
    )


# ── Sheet summary builder ──────────────────────────────────────────────────────

def describe_dataframe(df: pd.DataFrame, fname: str) -> str:
    """Return a human-readable summary string for LLM context."""
    key_cols, num_cols = classify_columns(df)
    lines = [
        f"파일: {fname}  ({len(df)}행 × {len(df.columns)}열)",
        f"  키 컬럼 : {', '.join(key_cols) or '없음'}",
        f"  숫자 컬럼: {', '.join(num_cols) or '없음'}",
        f"  샘플:",
    ]
    lines.append(df.head(3).to_string(index=False))
    return "\n".join(lines)
